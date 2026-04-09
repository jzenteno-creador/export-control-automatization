"""
data_extractor.py - Extractores de datos de documentos
Lee e interpreta PDFs, Excel y ZIPs automáticamente
"""

import logging
from pathlib import Path
from typing import Dict, Optional
import pypdf
import pdfplumber
import openpyxl
import json
from xml.etree import ElementTree as ET

logger = logging.getLogger(__name__)


class DataExtractor:
    """
    Extrae datos de documentos: PDFs, Excel, ZIPs
    """

    def __init__(self):
        logger.info("DataExtractor inicializado")

    # ========================
    # BILL OF LADING (BL)
    # ========================

    def extract_bl_data(self, pdf_path: str) -> Dict:
        """
        Extraer datos críticos de un Bill of Lading (PDF)
        
        Returns:
        {
            'producto': 'ACEITE VEGETAL',
            'cantidad': '100',
            'pesos': '50000',
            'contenedores': ['20FT-001', '40FT-002'],
            'precintos': ['ABC123', 'DEF456'],
            'consignee': 'IMPORTADOR USA',
            'notify': 'BROKER XYZ',
            'incoterm': 'CIF',
            'puerto_origen': 'BUENOS AIRES',
            'puerto_destino': 'NEW YORK',
            'linea_naviera': 'MAERSK'
        }
        """
        logger.info(f"Extrayendo datos de BL: {pdf_path}")
        
        try:
            # Leer PDF
            with pdfplumber.open(pdf_path) as pdf:
                text = ""
                tables = []
                
                # Extraer texto de todas las páginas
                for page in pdf.pages:
                    text += page.extract_text() or ""
                    # También extraer tablas
                    page_tables = page.extract_tables()
                    if page_tables:
                        tables.extend(page_tables)
                
                # Parsear datos
                data = self._parse_bl_text(text, tables)
                logger.info(f"✅ BL extraído: {data.get('consignee', 'UNKNOWN')}")
                return data
        
        except Exception as e:
            logger.error(f"Error extrayendo BL: {e}")
            return {'error': str(e)}

    def _parse_bl_text(self, text: str, tables: list) -> Dict:
        """
        Parsear texto y tablas de BL
        """
        data = {}
        
        # Buscar campos clave en el texto
        lines = text.split('\n')
        
        for i, line in enumerate(lines):
            line_upper = line.upper()
            
            # Consignee
            if 'CONSIGNEE' in line_upper or 'DESTINATARIO' in line_upper:
                data['consignee'] = self._extract_next_value(lines, i)
            
            # Notify
            if 'NOTIFY' in line_upper or 'NOTIFICAR' in line_upper:
                data['notify'] = self._extract_next_value(lines, i)
            
            # Producto
            if 'DESCRIPTION' in line_upper or 'DESCRIPCIÓN' in line_upper:
                data['producto'] = self._extract_next_value(lines, i)
            
            # Cantidad
            if 'QUANTITY' in line_upper or 'CANTIDAD' in line_upper:
                data['cantidad'] = self._extract_number(line)
            
            # Pesos
            if 'WEIGHT' in line_upper or 'PESO' in line_upper:
                data['pesos'] = self._extract_number(line)
            
            # Contenedores
            if 'CONTAINER' in line_upper or 'CONTENEDOR' in line_upper:
                data['contenedores'] = self._extract_containers(lines, i)
            
            # Precintos
            if 'SEAL' in line_upper or 'PRECINTO' in line_upper:
                data['precintos'] = self._extract_seals(lines, i)
            
            # Incoterm
            if 'INCOTERM' in line_upper:
                data['incoterm'] = self._extract_next_value(lines, i)
            
            # Puertos
            if 'PORT OF ORIGIN' in line_upper or 'PUERTO DE ORIGEN' in line_upper:
                data['puerto_origen'] = self._extract_next_value(lines, i)
            
            if 'PORT OF DESTINATION' in line_upper or 'PUERTO DE DESTINO' in line_upper:
                data['puerto_destino'] = self._extract_next_value(lines, i)
            
            # Línea naviera
            if 'CARRIER' in line_upper or 'NAVIERA' in line_upper:
                data['linea_naviera'] = self._extract_next_value(lines, i)
        
        return data

    # ========================
    # CERTIFICATE OF ORIGIN (COD)
    # ========================

    def extract_cod_data(self, xml_path: str) -> Dict:
        """
        Extraer datos de Certificado de Origen (XML de ZIP)
        
        Returns:
        {
            'producto': 'ACEITE VEGETAL',
            'pais_origen': 'ARGENTINA',
            'valor_declarado': '50000',
            'numero_certificado': 'CO-2024-001',
            'autoridad_emisora': 'CÁMARA DE COMERCIO'
        }
        """
        logger.info(f"Extrayendo datos de COD: {xml_path}")
        
        try:
            tree = ET.parse(xml_path)
            root = tree.getroot()
            
            data = {}
            
            # Buscar elementos en XML
            for elem in root.iter():
                tag = elem.tag.lower()
                
                if 'producto' in tag or 'description' in tag:
                    data['producto'] = elem.text
                elif 'pais' in tag or 'country' in tag:
                    data['pais_origen'] = elem.text
                elif 'valor' in tag or 'value' in tag:
                    data['valor_declarado'] = self._extract_number(elem.text or "")
                elif 'certificado' in tag or 'certificate' in tag:
                    data['numero_certificado'] = elem.text
                elif 'autoridad' in tag or 'authority' in tag:
                    data['autoridad_emisora'] = elem.text
            
            logger.info(f"✅ COD extraído: {data.get('numero_certificado', 'UNKNOWN')}")
            return data
        
        except Exception as e:
            logger.error(f"Error extrayendo COD: {e}")
            return {'error': str(e)}

    # ========================
    # BOOKING ADVICE
    # ========================

    def extract_booking_data(self, pdf_path: str) -> Dict:
        """
        Extraer datos de Booking Advice (PDF de SAP)
        
        Returns:
        {
            'consignee': 'IMPORTADOR USA',
            'notify': 'BROKER XYZ',
            'descripcion_cargo': 'ACEITE VEGETAL',
            'contenedores': ['20FT-001'],
            'cantidad': '100',
            'linea_naviera': 'MAERSK',
            'fecha_corte': '2024-03-30'
        }
        """
        logger.info(f"Extrayendo datos de Booking: {pdf_path}")
        
        try:
            with pdfplumber.open(pdf_path) as pdf:
                text = ""
                for page in pdf.pages:
                    text += page.extract_text() or ""
                
                data = self._parse_booking_text(text)
                logger.info(f"✅ Booking extraído: {data.get('consignee', 'UNKNOWN')}")
                return data
        
        except Exception as e:
            logger.error(f"Error extrayendo Booking: {e}")
            return {'error': str(e)}

    def _parse_booking_text(self, text: str) -> Dict:
        """
        Parsear texto de Booking Advice
        """
        data = {}
        lines = text.split('\n')
        
        for i, line in enumerate(lines):
            line_upper = line.upper()
            
            if 'CONSIGNEE' in line_upper:
                data['consignee'] = self._extract_next_value(lines, i)
            elif 'NOTIFY' in line_upper:
                data['notify'] = self._extract_next_value(lines, i)
            elif 'CARGO' in line_upper or 'DESCRIPCIÓN' in line_upper:
                data['descripcion_cargo'] = self._extract_next_value(lines, i)
            elif 'CONTENEDOR' in line_upper or 'CONTAINER' in line_upper:
                data['contenedores'] = self._extract_containers(lines, i)
            elif 'CARRIER' in line_upper or 'NAVIERA' in line_upper:
                data['linea_naviera'] = self._extract_next_value(lines, i)
        
        return data

    # ========================
    # PLANILLA DE ADUANA
    # ========================

    def extract_aduana_data(self, excel_path: str, sheet_name: str = 'Datos') -> Dict:
        """
        Extraer datos de Planilla de Aduana (Excel)
        
        Returns:
        {
            'contenedores': ['20FT-001', '40FT-002'],
            'precintos': ['ABC123', 'DEF456'],
            'pesos': '50000',
            'permiso_exportacion': 'EXP-2024-001',
            'cantidad_contenedores': '2',
            'clasificacion_arancelaria': '1507.10.00',
            'pais_destino': 'ESTADOS UNIDOS'
        }
        """
        logger.info(f"Extrayendo datos de Aduana: {excel_path}")
        
        try:
            wb = openpyxl.load_workbook(excel_path)
            
            # Encontrar la hoja correcta
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
            else:
                ws = wb.active
            
            data = {}
            
            # Recorrer celdas para encontrar datos
            for row in ws.iter_rows(values_only=True):
                for cell in row:
                    if cell is None:
                        continue
                    
                    cell_str = str(cell).upper()
                    
                    # Buscar campos por nombre
                    if 'CONTENEDOR' in cell_str:
                        data['contenedores'] = self._extract_from_row(row)
                    elif 'PRECINTO' in cell_str:
                        data['precintos'] = self._extract_from_row(row)
                    elif 'PESO' in cell_str:
                        data['pesos'] = self._extract_number(str(cell))
                    elif 'PERMISO' in cell_str or 'EXPORTACION' in cell_str:
                        data['permiso_exportacion'] = self._extract_from_row(row)
                    elif 'ARANCELARIA' in cell_str:
                        data['clasificacion_arancelaria'] = self._extract_from_row(row)
                    elif 'DESTINO' in cell_str:
                        data['pais_destino'] = self._extract_from_row(row)
            
            # Calcular cantidad de contenedores
            if 'contenedores' in data:
                data['cantidad_contenedores'] = str(len(data['contenedores']))
            
            logger.info(f"✅ Aduana extraída: {data.get('permiso_exportacion', 'UNKNOWN')}")
            return data
        
        except Exception as e:
            logger.error(f"Error extrayendo Aduana: {e}")
            return {'error': str(e)}

    # ========================
    # FACTURA COMERCIAL
    # ========================

    def extract_factura_data(self, excel_path: str, sheet_name: str = 'Factura') -> Dict:
        """
        Extraer datos de Factura Comercial (Excel o PDF)
        
        Returns:
        {
            'descripcion': 'ACEITE VEGETAL',
            'cantidad': '100',
            'valor_unitario': '500',
            'total': '50000',
            'moneda': 'USD',
            'incoterm': 'CIF',
            'vendedor': 'MI EMPRESA',
            'comprador': 'IMPORTADOR USA'
        }
        """
        logger.info(f"Extrayendo datos de Factura: {excel_path}")
        
        try:
            wb = openpyxl.load_workbook(excel_path)
            
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
            else:
                ws = wb.active
            
            data = {}
            
            # Recorrer celdas
            for row in ws.iter_rows(values_only=True):
                for cell in row:
                    if cell is None:
                        continue
                    
                    cell_str = str(cell).upper()
                    
                    if 'DESCRIPCION' in cell_str or 'DESCRIPTION' in cell_str:
                        data['descripcion'] = self._extract_from_row(row)
                    elif 'CANTIDAD' in cell_str or 'QUANTITY' in cell_str:
                        data['cantidad'] = self._extract_number(str(cell))
                    elif 'UNITARIO' in cell_str or 'UNIT' in cell_str:
                        data['valor_unitario'] = self._extract_number(str(cell))
                    elif 'TOTAL' in cell_str:
                        data['total'] = self._extract_number(str(cell))
                    elif 'MONEDA' in cell_str or 'CURRENCY' in cell_str:
                        data['moneda'] = self._extract_from_row(row)
                    elif 'INCOTERM' in cell_str:
                        data['incoterm'] = self._extract_from_row(row)
                    elif 'VENDEDOR' in cell_str or 'SELLER' in cell_str:
                        data['vendedor'] = self._extract_from_row(row)
                    elif 'COMPRADOR' in cell_str or 'BUYER' in cell_str:
                        data['comprador'] = self._extract_from_row(row)

            logger.info(f"Factura extraida: {data.get('descripcion', 'UNKNOWN')}")
            return data

        except Exception as e:
            logger.error(f"Error extrayendo Factura: {e}")
            return {'error': str(e)}

    # ========================
    # MÉTODOS AUXILIARES
    # ========================

    def _extract_next_value(self, lines: list, index: int) -> str:
        """Extraer el valor de la línea actual (después de ':') o la siguiente línea"""
        line = lines[index]
        if ':' in line:
            return line.split(':', 1)[1].strip()
        if index + 1 < len(lines):
            return lines[index + 1].strip()
        return ""

    def _extract_number(self, text: str) -> str:
        """Extraer número de un texto"""
        import re
        numbers = re.findall(r'[\d,.]+', text)
        if numbers:
            return numbers[0].replace(',', '')
        return ""

    def _extract_containers(self, lines: list, index: int) -> list:
        """Extraer lista de contenedores"""
        import re
        containers = []
        for line in lines[index:]:
            found = re.findall(r'[A-Z]{4}\d{7}', line)
            if found:
                containers.extend(found)
            elif containers:
                break
        return containers

    def _extract_seals(self, lines: list, index: int) -> list:
        """Extraer lista de precintos"""
        import re
        seals = []
        for line in lines[index:]:
            found = re.findall(r'[A-Z0-9]{5,}', line)
            if found:
                seals.extend(found)
            elif seals:
                break
        return seals

    def _extract_from_row(self, row: tuple) -> str:
        """Extraer valor significativo de una fila de Excel (primera celda no vacía después del label)"""
        values = [str(cell) for cell in row if cell is not None]
        if len(values) > 1:
            return values[1].strip()
        return values[0].strip() if values else ""